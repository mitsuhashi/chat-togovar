#!/usr/bin/env python

import os
import re
import json5
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# 設定
directory_path = "./evaluation/gpt-4o"
questions_path = "questions.json"
github_base_url_evaluation = "https://github.com/mitsuhashi/chat-togovar/blob/main/evaluation/gpt-4o"
model_link_bases = {
    "ChatTogoVar": "https://github.com/mitsuhashi/chat-togovar/blob/main/answers/chat_togovar",
    "GPT-4o": "https://github.com/mitsuhashi/chat-togovar/blob/main/answers/gpt-4o",
    "VarChat": "https://github.com/mitsuhashi/chat-togovar/blob/main/answers/varchat"
}
models = ["ChatTogoVar", "GPT-4o", "VarChat"]
categories = ["Accuracy", "Completeness", "Logical Consistency", "Clarity and Conciseness", "Evidence Support"]
category_colors = {
    "Accuracy": "B7DEE8", "Completeness": "D9EAD3", "Logical Consistency": "F9CB9C",
    "Clarity and Conciseness": "EAD1DC", "Evidence Support": "FFF2CC"
}
score_patterns = {
    "BestAnswer": r"Best Answer: (.+)",
    "ChatTogoVar": r"Total Score for ChatTogoVar: (\d+)/\d+",
    "GPT-4o": r"Total Score for GPT-4o: (\d+)/\d+",
    "VarChat": r"Total Score for VarChat: (\d+)/\d+"
}

def load_questions(path):
    with open(path, "r") as f:
        return json5.load(f)

def parse_markdown_file(filepath, question_no, question_template):
    rs = os.path.basename(filepath).replace(".md", "")
    with open(filepath, encoding='utf-8') as f:
        content = f.read()

    record = {
        "QuestionNumber": question_no,
        "Question": question_template.format(rs=rs),
        "Filename_Text": f"{rs}.md"
    }

    for key, pattern in score_patterns.items():
        match = re.search(pattern, content)
        record[key] = match.group(1) if match else None

    for model in models:
        block_pattern = rf"## Answer {re.escape(model)}\n(.*?)(?:\n## Answer |\Z)"
        match = re.search(block_pattern, content, re.DOTALL)
        if match:
            block = match.group(1)
            for cat in categories:
                cat_pattern = rf"- {cat} Score: (\d+)/10"
                cat_match = re.search(cat_pattern, block)
                record[f"{model}_{cat}"] = int(cat_match.group(1)) if cat_match else None
        else:
            for cat in categories:
                record[f"{model}_{cat}"] = None
    return record

def calculate_win_rates(df):
    counts = df["BestAnswer"].value_counts().reindex(models, fill_value=0)
    return pd.DataFrame({
        "Model": counts.index,
        "Wins": counts.values,
        "Win Rate (%)": (counts.values / len(df) * 100).round(2)
    })

def calculate_category_averages(df):
    records = []
    for criterion in categories:
        record = {"Criterion": criterion}
        for model in models:
            col = f"{model}_{criterion}"
            if col in df.columns:
                record[model] = df[col].mean()
        records.append(record)
    return pd.DataFrame(records)

def export_to_excel(df, winrate_df, category_df, output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # メインシート
        df.to_excel(writer, sheet_name="Sheet1", index=False)
        workbook = writer.book
        sheet = workbook["Sheet1"]
        sheet.auto_filter.ref = sheet.dimensions

        for row in range(2, sheet.max_row + 1):
            filename = sheet.cell(row=row, column=df.columns.get_loc("Filename_Text") + 1).value
            q_no = sheet.cell(row=row, column=df.columns.get_loc("QuestionNumber") + 1).value
            url = f"{github_base_url_evaluation}/{q_no}/{filename}"
            sheet.cell(row=row, column=df.columns.get_loc("Filename_Text") + 1).hyperlink = url
            sheet.cell(row=row, column=df.columns.get_loc("Filename_Text") + 1).style = "Hyperlink"

            for model in models:
                model_col = df.columns.get_loc(model) + 1 if model in df.columns else None
                if model_col:
                    model_url = f"{model_link_bases[model]}/{q_no}/{filename}" if model != "VarChat" else f"{model_link_bases[model]}/{filename}"
                    sheet.cell(row=row, column=model_col).hyperlink = model_url
                    sheet.cell(row=row, column=model_col).style = "Hyperlink"

        for model in models:
            for cat in categories:
                col = f"{model}_{cat}"
                if col in df.columns:
                    col_idx = df.columns.get_loc(col) + 1
                    fill = PatternFill(start_color=category_colors[cat], end_color=category_colors[cat], fill_type="solid")
                    for row in range(2, sheet.max_row + 1):
                        sheet.cell(row=row, column=col_idx).fill = fill
                    sheet.column_dimensions[get_column_letter(col_idx)].width = 5

        # 勝率シート
        winrate_df.to_excel(writer, sheet_name="Win Rates", index=False)

        # カテゴリ平均シート
        category_df.to_excel(writer, sheet_name="Category Averages", index=False)
        chart_sheet = writer.sheets["Category Averages"]

        chart = BarChart()
        chart.title = "Average Scores per Category"
        chart.x_axis.title = "Criterion"
        chart.y_axis.title = "Average Score"
        chart.style = 11

        data = Reference(chart_sheet, min_col=2, max_col=1 + len(models),
                        min_row=1, max_row=1 + len(category_df))
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=1 + len(category_df))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart_sheet.add_chart(chart, "E2")

def export_to_json(df, path):
    df.to_json(path, orient="records", force_ascii=False, indent=4)

def main():
    questions = load_questions(questions_path)
    df = pd.DataFrame()

    for q_no, template in questions.items():
        q_dir = os.path.join(directory_path, q_no)
        if not os.path.isdir(q_dir):
            continue
        for fname in os.listdir(q_dir):
            if fname.endswith(".md"):
                filepath = os.path.join(q_dir, fname)
                rec = parse_markdown_file(filepath, q_no, template)
                df = pd.concat([df, pd.DataFrame([rec])], ignore_index=True)

    for col in df.columns:
        if any(model in col for model in models):
            df[col] = pd.to_numeric(df[col], errors='coerce').astype("Int64")

    winrate_df = calculate_win_rates(df)
    category_df = calculate_category_averages(df)

    output_xlsx = os.path.join(directory_path, "aggregate_aqes.xlsx")
    export_to_excel(df, winrate_df, category_df, output_xlsx)

    output_json = os.path.join(directory_path, "aggregate_aqes.json")
    export_to_json(df, output_json)

    print(f"{output_xlsx} にExcel出力しました")
    print(f"{output_json} にJSON出力しました")

if __name__ == "__main__":
    main()